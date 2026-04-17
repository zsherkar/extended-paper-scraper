"""Export a user-friendly distillation paper database and optional Google Sheet bundle."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import scripts.extract_distillation as distill


DEFAULT_OUTPUT_PREFIX = Path("outputs/distillation_paper_database")
DEFAULT_GOOGLE_SHEET_ID = "1boGyiTko5h50pHqHZQK64X4aO4iK3U6pDlSnzh2F0B4"
EXCEL_CELL_LIMIT = 32767

DATABASE_FIELDNAMES = [
    "paper_id",
    "title",
    "authors",
    "year",
    "venue",
    "conference_id",
    "citation_count",
    "score",
    "priority_band",
    "matched_group_count",
    "matched_term_count",
    "groups",
    "matched_terms",
    "primary_group",
    "is_llm",
    "is_attack",
    "is_black_box",
    "is_defense",
    "is_privacy",
    "is_compression",
    "is_data_synthetic",
    "is_policy",
    "is_advanced",
    "has_citations",
    "link",
    "tldr",
    "abstract",
]

SUMMARY_FIELDNAMES = ["metric", "value"]
GROUP_COUNT_FIELDNAMES = ["group", "paper_count"]
GOOGLE_SHEETS_INSTRUCTIONS = """
Google Sheets sync requires:
1. A service account JSON key.
2. The service account email added as an editor to the target sheet.
3. Optional packages installed: google-api-python-client and google-auth.
""".strip()


def _priority_band(score: int, citations: int) -> str:
    if score >= 120 or citations >= 300:
        return "A_high_signal"
    if score >= 60 or citations >= 50:
        return "B_strong"
    if score >= 30 or citations >= 10:
        return "C_relevant"
    return "D_broad"


def _paper_id(row: dict) -> str:
    key = "|".join(
        [
            str(row.get("title", "")).strip().casefold(),
            str(row.get("authors", "")).strip().casefold(),
            str(row.get("conference_id", "")).strip().casefold(),
        ]
    )
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


def _flag(groups: set[str], *names: str) -> str:
    return "TRUE" if any(name in groups for name in names) else "FALSE"


def _database_row(row: dict) -> dict:
    groups = set(row["groups"])
    matched_terms = row["matched_terms"]
    citations = int(row.get("citation_count") or 0)
    primary_group = row["groups"][0] if row["groups"] else ""
    return {
        "paper_id": _paper_id(row),
        "title": row["title"],
        "authors": row["authors"],
        "year": row["year"],
        "venue": row["venue"],
        "conference_id": row["conference_id"],
        "citation_count": citations,
        "score": row["score"],
        "priority_band": _priority_band(row["score"], citations),
        "matched_group_count": len(groups),
        "matched_term_count": len(matched_terms),
        "groups": "; ".join(row["groups"]),
        "matched_terms": "; ".join(matched_terms),
        "primary_group": primary_group,
        "is_llm": _flag(groups, "llm", "llm_foundation"),
        "is_attack": _flag(groups, "attack", "extraction_stealing", "attack_mechanism", "adversarial_ml"),
        "is_black_box": _flag(groups, "black_box", "query_black_box"),
        "is_defense": _flag(groups, "defense", "defense_extraction", "watermarking_ip"),
        "is_privacy": _flag(groups, "leakage", "privacy_leakage"),
        "is_compression": _flag(groups, "compression", "compression_overlap"),
        "is_data_synthetic": _flag(groups, "data", "data_free_synthetic"),
        "is_policy": _flag(groups, "policy", "economic_system"),
        "is_advanced": _flag(groups, "advanced", "advanced_niche"),
        "has_citations": "TRUE" if citations > 0 else "FALSE",
        "link": row["link"],
        "tldr": row["tldr"],
        "abstract": row["abstract"],
    }


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _excel_value(value):
    text = "" if value is None else str(value)
    if len(text) > EXCEL_CELL_LIMIT:
        return text[: EXCEL_CELL_LIMIT - 3] + "..."
    return text


def _append_sheet(workbook: Workbook, title: str, rows: list[dict], fieldnames: list[str]) -> None:
    sheet = workbook.create_sheet(title[:31])
    sheet.append(fieldnames)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        sheet.append([_excel_value(row.get(field, "")) for field in fieldnames])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        column_letter = column[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column[:100])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)


def _write_xlsx(path: Path, database_rows: list[dict], summary_rows: list[dict], group_rows: list[dict]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _append_sheet(workbook, "papers_database", database_rows, DATABASE_FIELDNAMES)
    _append_sheet(workbook, "summary", summary_rows, SUMMARY_FIELDNAMES)
    _append_sheet(workbook, "group_counts", group_rows, GROUP_COUNT_FIELDNAMES)
    _append_sheet(workbook, "web_sources", distill.WEB_SOURCES, distill.WEB_SOURCE_FIELDNAMES)
    _append_sheet(workbook, "query_taxonomy", distill._taxonomy_rows(), ["group", "term", "regex", "weight"])
    workbook.save(path)


def _build_summary(database_rows: list[dict], scanned: int, min_score: int) -> tuple[list[dict], list[dict]]:
    counts = Counter()
    by_year = Counter()
    by_venue = Counter()
    priority = Counter()

    for row in database_rows:
        groups = [item.strip() for item in row["groups"].split(";") if item.strip()]
        counts.update(groups)
        if row["year"]:
            by_year[row["year"]] += 1
        if row["venue"]:
            by_venue[row["venue"]] += 1
        priority[row["priority_band"]] += 1

    summary_rows = [
        {"metric": "papers_scanned", "value": scanned},
        {"metric": "candidate_papers", "value": len(database_rows)},
        {"metric": "minimum_score", "value": min_score},
        {"metric": "top_year", "value": by_year.most_common(1)[0][0] if by_year else ""},
        {"metric": "top_venue", "value": by_venue.most_common(1)[0][0] if by_venue else ""},
        {"metric": "high_signal_count", "value": priority.get("A_high_signal", 0)},
        {"metric": "strong_count", "value": priority.get("B_strong", 0)},
        {"metric": "google_sheet_id", "value": DEFAULT_GOOGLE_SHEET_ID},
    ]
    group_rows = [{"group": group, "paper_count": count} for group, count in counts.most_common()]
    return summary_rows, group_rows


def _write_google_sheet_bundle(output_dir: Path, database_rows: list[dict], summary_rows: list[dict], group_rows: list[dict]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    _write_csv(output_dir / "papers_database.csv", database_rows, DATABASE_FIELDNAMES)
    _write_csv(output_dir / "summary.csv", summary_rows, SUMMARY_FIELDNAMES)
    _write_csv(output_dir / "group_counts.csv", group_rows, GROUP_COUNT_FIELDNAMES)
    _write_csv(output_dir / "web_sources.csv", distill.WEB_SOURCES, distill.WEB_SOURCE_FIELDNAMES)
    _write_csv(output_dir / "query_taxonomy.csv", distill._taxonomy_rows(), ["group", "term", "regex", "weight"])
    (output_dir / "README.txt").write_text(GOOGLE_SHEETS_INSTRUCTIONS + "\n", encoding="utf-8")


def _sync_google_sheet(sheet_id: str, credentials_path: Path, database_rows: list[dict], summary_rows: list[dict], group_rows: list[dict]) -> None:
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Google Sheets sync requires `google-auth` and `google-api-python-client`."
        ) from exc

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    credentials = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
    service = build("sheets", "v4", credentials=credentials)

    def values(rows: list[dict], fieldnames: list[str]) -> list[list[str]]:
        header = [field for field in fieldnames]
        body = [[str(row.get(field, "")) for field in fieldnames] for row in rows]
        return [header, *body]

    requests = [
        {"addSheet": {"properties": {"title": title}}}
        for title in ["papers_database", "summary", "group_counts", "web_sources", "query_taxonomy"]
    ]
    try:
        service.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id,
            body={"requests": requests},
        ).execute()
    except Exception:
        # Tabs may already exist; keep going and overwrite contents.
        pass

    value_payloads = [
        ("papers_database!A1", values(database_rows, DATABASE_FIELDNAMES)),
        ("summary!A1", values(summary_rows, SUMMARY_FIELDNAMES)),
        ("group_counts!A1", values(group_rows, GROUP_COUNT_FIELDNAMES)),
        ("web_sources!A1", values(distill.WEB_SOURCES, distill.WEB_SOURCE_FIELDNAMES)),
        ("query_taxonomy!A1", values(distill._taxonomy_rows(), ["group", "term", "regex", "weight"])),
    ]
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=sheet_id,
        body={
            "valueInputOption": "RAW",
            "data": [{"range": range_name, "values": vals} for range_name, vals in value_payloads],
        },
    ).execute()


def main() -> None:
    parser = argparse.ArgumentParser(description="Export a filterable distillation paper database.")
    parser.add_argument("--data-dir", type=Path, default=distill.DEFAULT_DATA_DIR)
    parser.add_argument("--master-csv", type=Path, default=Path("master_literature.csv"))
    parser.add_argument("--no-master-csv", action="store_true", help="Skip master_literature.csv even if present.")
    parser.add_argument("--output-prefix", type=Path, default=DEFAULT_OUTPUT_PREFIX)
    parser.add_argument("--min-score", type=int, default=12)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--google-sheet-id", default=DEFAULT_GOOGLE_SHEET_ID)
    parser.add_argument("--google-credentials", type=Path, help="Path to a Google service account JSON key.")
    parser.add_argument("--sync-google-sheet", action="store_true", help="Push the generated tabs into Google Sheets.")
    args = parser.parse_args()

    master_csv = None if args.no_master_csv else args.master_csv
    rows, scanned = distill.extract(args.data_dir, min_score=args.min_score, master_csv=master_csv)
    if args.limit:
        rows = rows[: args.limit]

    database_rows = [_database_row(row) for row in rows]
    summary_rows, group_rows = _build_summary(database_rows, scanned=scanned, min_score=args.min_score)

    args.output_prefix.parent.mkdir(parents=True, exist_ok=True)
    _write_csv(args.output_prefix.with_suffix(".csv"), database_rows, DATABASE_FIELDNAMES)
    with open(args.output_prefix.with_suffix(".jsonl"), "w", encoding="utf-8") as f:
        for row in database_rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    _write_xlsx(args.output_prefix.with_suffix(".xlsx"), database_rows, summary_rows, group_rows)
    _write_google_sheet_bundle(args.output_prefix.with_name(f"{args.output_prefix.name}_google_sheet"), database_rows, summary_rows, group_rows)

    if args.sync_google_sheet:
        if args.google_credentials is None:
            raise RuntimeError("`--sync-google-sheet` requires `--google-credentials`.")
        _sync_google_sheet(args.google_sheet_id, args.google_credentials, database_rows, summary_rows, group_rows)

    print(f"Scanned {scanned} papers")
    print(f"Wrote {len(database_rows)} database rows to {args.output_prefix.with_suffix('.csv')}")
    print(f"Excel workbook: {args.output_prefix.with_suffix('.xlsx')}")
    print(f"Google Sheet bundle: {args.output_prefix.with_name(f'{args.output_prefix.name}_google_sheet')}")
    if args.sync_google_sheet:
        print(f"Google Sheet synced: https://docs.google.com/spreadsheets/d/{args.google_sheet_id}/edit")
    else:
        print(f"Google Sheet target: https://docs.google.com/spreadsheets/d/{args.google_sheet_id}/edit")
        if args.google_credentials is None:
            print("Google Sheet sync skipped: no service account credentials provided.")


if __name__ == "__main__":
    main()
