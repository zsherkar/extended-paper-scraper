"""Extract AI/ML distillation-related papers from local paper JSONL data."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_DATA_DIR = Path("data")
DEFAULT_OUTPUT_PREFIX = Path("outputs/distillation_candidates")
EXCEL_CELL_LIMIT = 32767
FIELDNAMES = [
    "score",
    "groups",
    "matched_terms",
    "conference_id",
    "venue",
    "year",
    "title",
    "authors",
    "link",
    "citation_count",
    "tldr",
    "abstract",
]
WEB_SOURCE_FIELDNAMES = [
    "source_type",
    "category",
    "title",
    "url",
    "notes",
]

FIELD_WEIGHTS = {
    "title": 5,
    "keywords": 4,
    "tldr": 3,
    "abstract": 2,
}

PATTERN_GROUPS: dict[str, list[tuple[str, str, int]]] = {
    "distillation": [
        ("knowledge distillation", r"\bknowledge distillation\b", 6),
        ("model distillation", r"\bmodel\s+distillation\b", 6),
        ("neural network distillation", r"\bneural\s+network\s+distillation\b", 6),
        ("distillation", r"\bdistill(?:ation|ed|ing|s)?\b", 5),
        ("teacher-student", r"\bteacher[-\s]?student\b", 5),
        ("teacher-student framework", r"\bteacher[-\s]?student\s+framework\b", 6),
        ("teacher-student learning", r"\bteacher[-\s]?student\s+learning\b", 6),
        ("teacher model", r"\bteacher model\b", 3),
        ("student model", r"\bstudent model\b", 3),
        ("student model training", r"\bstudent\s+model\s+training\b", 4),
        ("soft label distillation", r"\bsoft[-\s]?labels?\s+distillation\b", 5),
        ("hard label distillation", r"\bhard[-\s]?labels?\s+distillation\b", 5),
        ("temperature scaling distillation", r"\btemperature\s+scal(?:ing|ed).{0,25}distillation\b", 5),
        ("self-distillation", r"\bself[-\s]?distillation\b", 6),
        ("logit distillation", r"\blogits?\s+distillation\b|\bdistillation\s+logits?\b", 5),
        ("feature distillation", r"\bfeature\s+distillation\b", 5),
        ("representation distillation", r"\brepresentation\s+distillation\b", 5),
        ("intermediate layer distillation", r"\bintermediate\s+layers?\s+distillation\b", 5),
        ("hint-based distillation", r"\bhint[-\s]?based\s+distillation\b|\bFitNets?\b", 5),
        ("attention transfer", r"\battention\s+transfer\b", 4),
        ("relational knowledge distillation", r"\brelational\s+knowledge\s+distillation\b", 6),
        ("distillation loss", r"\bdistillation\s+loss\b", 4),
        ("KL divergence distillation", r"\bKL\s+divergence.{0,25}distill", 4),
        ("cross entropy distillation", r"\bcross[-\s]?entropy.{0,25}distill", 4),
        ("data distillation", r"\bdata\s+distillation\b", 6),
        ("dataset distillation", r"\bdataset\s+distillation\b", 6),
        ("policy distillation", r"\bpolicy\s+distillation\b", 5),
        ("trajectory distillation", r"\btrajectory\s+distillation\b", 5),
        ("response distillation", r"\bresponse\s+distillation\b", 5),
        ("rationale distillation", r"\brationale\s+distillation\b", 5),
    ],
    "distillation_type": [
        ("offline distillation", r"\boffline\s+distillation\b", 5),
        ("online distillation", r"\bonline\s+distillation\b", 5),
        ("mutual distillation", r"\bmutual\s+distillation\b", 5),
        ("peer distillation", r"\bpeer\s+distillation\b", 5),
        ("sequence distillation", r"\bsequence\s+distillation\b", 5),
        ("sequence-level distillation", r"\bsequence[-\s]?level\s+distillation\b", 5),
        ("token-level distillation", r"\btoken[-\s]?level\s+distillation\b", 5),
        ("structured prediction distillation", r"\bstructured\s+prediction\s+distillation\b", 5),
        ("transformer distillation", r"\btransformer\s+distillation\b", 5),
        ("BERT distillation", r"\bBERT\s+distillation\b|\bDistilBERT\b", 5),
        ("GPT distillation", r"\bGPT[-\w]*\s+distillation\b", 5),
        ("diffusion model distillation", r"\bdiffusion\s+models?\s+distillation\b|\bdistillation.{0,30}diffusion\s+models?\b", 5),
        ("vision transformer distillation", r"\bvision\s+transformer\s+distillation\b|\bViT\s+distillation\b", 5),
        ("CNN distillation", r"\bCNN\s+distillation\b|\bconvolutional\s+neural\s+network\s+distillation\b", 4),
    ],
    "compression": [
        ("model compression", r"\bmodel\s+compression\b", 4),
        ("neural compression", r"\bneural\s+compression\b", 3),
        ("pruning and distillation", r"\b(?:pruning|prune).{0,40}distill|\bdistill.{0,40}(?:pruning|prune)\b", 5),
        ("quantization-aware distillation", r"\bquantization[-\s]?aware\s+distillation\b|\bdistill.{0,40}quantization\b", 5),
        ("low-rank distillation", r"\blow[-\s]?rank\s+distillation\b", 4),
        ("parameter-efficient distillation", r"\bparameter[-\s]?efficient\s+distillation\b", 5),
        ("edge distillation", r"\bdistillation\s+for\s+edge\s+devices\b|\bedge.{0,30}distillation\b", 4),
        ("small model efficiency", r"\b(?:tiny|small|compact)\s+models?.{0,40}(?:distill|efficien)", 3),
    ],
    "data": [
        ("data-free distillation", r"\bdata[-\s]?free\s+distillation\b", 6),
        ("dataset-free distillation", r"\bdataset[-\s]?free\s+distillation\b", 6),
        ("zero-shot distillation", r"\bzero[-\s]?shot\s+distillation\b", 6),
        ("synthetic data distillation", r"\bsynthetic\s+data\s+distillation\b", 6),
        ("generative distillation", r"\bgenerative\s+distillation\b", 5),
        ("pseudo-labeling", r"\bpseudo[-\s]?labell?ing\b", 2),
        ("unlabeled data distillation", r"\bunlabeled\s+data\s+distillation\b", 5),
        ("data-free", r"\bdata[-\s]?free\b", 3),
        ("synthetic queries", r"\bsynthetic\s+quer(?:y|ies)\b", 4),
    ],
    "llm": [
        ("large language model", r"\blarge language models?\b", 3),
        ("LLM", r"\bLLMs?\b", 3),
        ("language model", r"\blanguage models?\b", 2),
        ("GPT", r"\bGPT[-\w]*\b", 2),
        ("LLM distillation", r"\bLLM\s+distillation\b|\blarge\s+language\s+models?\s+distillation\b|\bdistill(?:ation|ed|ing)?.{0,40}large\s+language\s+models?\b", 6),
        ("instruction distillation", r"\binstruction\s+distillation\b", 6),
        ("alignment distillation", r"\balignment\s+distillation\b", 6),
        ("RLHF distillation", r"\bRLHF\s+distillation\b|\breinforcement\s+learning\s+from\s+human\s+feedback.{0,40}distill", 6),
        ("reasoning distillation", r"\breasoning\s+distillation\b", 6),
        ("chain-of-thought distillation", r"\bchain[-\s]of[-\s]thought\s+distillation\b", 6),
        ("preference distillation", r"\bpreference\s+distillation\b", 5),
        ("reward model distillation", r"\breward\s+model\s+distillation\b", 5),
        ("self-improving distillation", r"\bself[-\s]?improving\s+models?.{0,40}distill", 4),
    ],
    "attack": [
        ("model extraction", r"\bmodel\s+extraction\b", 7),
        ("model stealing", r"\bmodel\s+steal(?:ing)?\b", 7),
        ("neural network stealing", r"\bneural\s+network\s+steal(?:ing)?\b", 7),
        ("model replication attack", r"\bmodel\s+replication\s+attack\b", 7),
        ("model cloning", r"\bmodel\s+clon(?:e|ing)\b", 6),
        ("stealing attack", r"\bsteal(?:ing)?\s+attack\b", 6),
        ("extraction attack", r"\bextraction\s+attack\b", 7),
        ("black-box extraction", r"\bblack[-\s]?box\s+extraction\b", 6),
        ("API-based model extraction", r"\bAPI[-\s]?based\s+model\s+extraction\b|\bmodel\s+extraction.{0,40}API\b", 7),
        ("extraction via distillation", r"\bextraction\s+via\s+distillation\b", 7),
        ("distillation-based model stealing", r"\bdistillation[-\s]?based\s+model\s+steal(?:ing)?\b", 7),
        ("student-teacher attack", r"\bstudent[-\s]?teacher\s+attack\b", 6),
        ("query-based distillation attack", r"\bquery[-\s]?based\s+distillation\s+attack\b", 7),
        ("knockoff", r"\bknockoff\b", 5),
        ("copycat", r"\bcopycat\b", 5),
    ],
    "black_box": [
        ("black-box attack", r"\bblack[-\s]?box\s+attack\b", 5),
        ("query-efficient model extraction", r"\bquery[-\s]?efficient\s+model\s+extraction\b", 7),
        ("adaptive querying", r"\badaptive\s+quer(?:y|ying|ies)\b", 4),
        ("active learning for extraction", r"\bactive\s+learning.{0,40}extraction\b|\bextraction.{0,40}active\s+learning\b", 4),
        ("query synthesis attack", r"\bquery\s+synthesis\s+attack\b", 6),
        ("decision-based extraction", r"\bdecision[-\s]?based\s+extraction\b", 6),
        ("score-based extraction", r"\bscore[-\s]?based\s+extraction\b", 6),
    ],
    "leakage": [
        ("training data extraction", r"\btraining\s+data\s+extraction\b", 6),
        ("dataset reconstruction attack", r"\bdataset\s+reconstruction\s+attack\b", 6),
        ("data leakage via distillation", r"\bdata\s+leakage.{0,40}distill", 6),
        ("inversion attack", r"\binversion\s+attacks?\b", 4),
        ("model inversion via distillation", r"\bmodel\s+inversion.{0,40}distill", 6),
        ("membership inference distillation", r"\bmembership\s+inference.{0,40}distill|\bdistill.{0,40}membership\s+inference\b", 5),
        ("gradient leakage", r"\bgradient\s+leak(?:age|s)?\b", 4),
    ],
    "defense": [
        ("model watermarking", r"\bmodel\s+watermark(?:ing)?\b", 6),
        ("neural watermarking", r"\bneural\s+watermark(?:ing)?\b", 5),
        ("distillation-resistant watermarking", r"\bdistillation[-\s]?resistant\s+watermark(?:ing)?\b", 7),
        ("watermark robustness to distillation", r"\bwatermark.{0,40}robust.{0,40}distill|\bdistill.{0,40}watermark", 6),
        ("model fingerprinting", r"\bmodel\s+fingerprint(?:ing)?\b", 5),
        ("ownership verification", r"\bownership\s+verification\b", 5),
        ("model provenance", r"\bmodel\s+provenance\b|\bprovenance\s+tracking\b", 5),
        ("IP protection", r"\bIP\s+protection\b|\bintellectual\s+property\s+protection\b", 4),
        ("extraction defense", r"\bextraction\s+defen[cs]e\b", 6),
        ("model stealing defense", r"\bmodel\s+steal(?:ing)?\s+defen[cs]e\b", 6),
        ("query monitoring", r"\bquery\s+monitoring\b", 5),
        ("rate limiting", r"\brate\s+limit(?:ing)?\b", 3),
        ("output perturbation", r"\boutput\s+perturbation\b", 5),
        ("prediction obfuscation", r"\bprediction\s+obfuscation\b", 5),
        ("confidence masking", r"\bconfidence\s+masking\b", 5),
        ("adversarial response", r"\badversarial\s+response\s+strateg(?:y|ies)\b", 5),
        ("anti-distillation", r"\banti[-\s]?distillation\b", 6),
    ],
    "technique": [
        ("zero-shot", r"\bzero[-\s]?shot\b", 2),
        ("contrastive distillation", r"\bcontrastive\s+distillation\b", 5),
        ("adversarial distillation", r"\badversarial\s+distillation\b", 5),
        ("defensive distillation", r"\bdefensive\s+distillation\b", 6),
        ("distillation for robustness", r"\bdistillation\s+for\s+robustness\b|\brobustness.{0,30}distillation\b", 4),
        ("adversarial training distillation", r"\badversarial\s+training.{0,25}distillation\b|\bdistillation.{0,25}adversarial\s+training\b", 5),
        ("transfer attacks via distillation", r"\btransfer\s+attacks?.{0,40}distill|\bdistill.{0,40}transfer\s+attacks?\b", 5),
        ("federated distillation", r"\bfederated\s+distillation\b", 5),
        ("ensemble distillation", r"\bensemble\s+distillation\b", 4),
        ("cross-modal distillation", r"\bcross[-\s]?modal\s+distillation\b", 4),
        ("surrogate model", r"\bsurrogate\s+model\b", 2),
        ("substitute model", r"\bsubstitute\s+model\b", 2),
    ],
    "policy": [
        ("API abuse", r"\bAPI\s+abuse\b", 5),
        ("MLaaS security", r"\bMLaaS\s+security\b|\bmachine\s+learning\s+as\s+a\s+service\b", 5),
        ("cloud inference security", r"\bcloud\s+inference\s+security\b", 4),
        ("proprietary model protection", r"\bproprietary\s+model\s+protection\b", 5),
        ("compute asymmetry", r"\bcompute\s+asymmetry\b", 5),
        ("training vs extraction cost", r"\bcost\s+of\s+model\s+training.{0,50}extraction\b|\btraining\s+cost.{0,50}extraction\b", 5),
        ("AI model IP law", r"\bAI\s+model\s+IP\s+law\b|\bmodel\s+IP\s+law\b", 4),
        ("export controls", r"\b(?:AI|model|compute|chip|GPU|semiconductor).{0,40}export\s+controls?\b|\bexport\s+controls?.{0,40}(?:AI|model|compute|chip|GPU|semiconductor)\b", 4),
        ("dual use AI", r"\bdual[-\s]?use\s+AI\b|\bdual[-\s]?use\s+artificial\s+intelligence\b", 4),
        ("AI supply chain security", r"\bAI\s+supply\s+chain\s+security\b", 4),
    ],
    "advanced": [
        ("distillation collapse", r"\bdistillation\s+collapse\b", 5),
        ("student-teacher mismatch", r"\bstudent[-\s]?teacher\s+mismatch\b", 5),
        ("capacity gap", r"\bcapacity\s+gap.{0,30}distill|\bdistill.{0,30}capacity\s+gap\b", 4),
        ("iterative distillation", r"\biterative\s+distillation\b", 5),
        ("recursive distillation", r"\brecursive\s+distillation\b", 5),
        ("self-improving distillation loops", r"\bself[-\s]?improving\s+distillation\s+loops?\b", 5),
        ("distillation stability", r"\bdistillation\s+stability\b|\bstability.{0,30}distillation\b", 4),
        ("distillation scaling laws", r"\bdistillation\s+scaling\s+laws?\b|\bscaling\s+laws?.{0,30}distillation\b", 5),
        ("knowledge transfer efficiency", r"\bknowledge\s+transfer\s+efficiency\b", 4),
    ],
}

EXHAUSTIVE_KEYWORD_GROUPS: dict[str, list[str]] = {
    "core_foundational": [
        "knowledge distillation",
        "model distillation",
        "neural network distillation",
        "teacher student learning",
        "teacher student model",
        "student model training",
        "teacher model outputs",
        "soft label distillation",
        "hard label distillation",
        "logits distillation",
        "temperature scaling distillation",
        "distillation loss",
        "KL divergence distillation",
        "cross entropy distillation",
        "dark knowledge distillation",
    ],
    "representation_feature": [
        "feature distillation",
        "representation distillation",
        "hidden layer distillation",
        "intermediate layer distillation",
        "hint-based distillation",
        "FitNets",
        "attention transfer",
        "activation matching",
        "embedding distillation",
        "relational knowledge distillation",
        "similarity preserving distillation",
    ],
    "framework_variant": [
        "offline distillation",
        "online distillation",
        "self distillation",
        "mutual distillation",
        "peer distillation",
        "collaborative distillation",
        "ensemble distillation",
        "multi teacher distillation",
        "teacher ensemble compression",
    ],
    "task_output": [
        "sequence distillation",
        "sequence level distillation",
        "token level distillation",
        "structured prediction distillation",
        "policy distillation",
        "value function distillation",
        "imitation learning distillation",
        "behavior cloning distillation",
    ],
    "llm_foundation": [
        "LLM distillation",
        "large language model distillation",
        "transformer distillation",
        "BERT distillation",
        "GPT distillation",
        "instruction distillation",
        "alignment distillation",
        "RLHF distillation",
        "reward model distillation",
        "policy distillation RL",
        "chain of thought distillation",
        "reasoning distillation",
        "self improvement distillation",
        "recursive distillation",
    ],
    "data_free_synthetic": [
        "data free distillation",
        "dataset free distillation",
        "zero shot distillation",
        "few shot distillation",
        "synthetic data distillation",
        "generative distillation",
        "pseudo label distillation",
        "distillation without data",
        "model inversion distillation",
        "data reconstruction distillation",
    ],
    "compression_overlap": [
        "model compression distillation",
        "network compression distillation",
        "model pruning distillation",
        "quantization distillation",
        "low rank distillation",
        "parameter efficient distillation",
        "efficient student models",
        "tiny models distillation",
        "edge deployment distillation",
    ],
    "extraction_stealing": [
        "model extraction attack",
        "model stealing",
        "neural network stealing",
        "model cloning",
        "model replication",
        "black box model extraction",
        "API model extraction",
        "distillation based extraction",
        "distillation based model stealing",
        "student teacher attack",
        "surrogate model extraction",
    ],
    "query_black_box": [
        "black box attack machine learning",
        "query based model extraction",
        "query efficient extraction",
        "adaptive query attack",
        "active learning attack",
        "decision based attack",
        "score based attack",
        "confidence based extraction",
        "label only extraction",
        "probability output attack",
    ],
    "attack_mechanism": [
        "adversarial distillation",
        "malicious distillation",
        "distillation attack",
        "knowledge extraction via distillation",
        "distillation based reverse engineering",
        "distillation driven model copying",
        "distillation based surrogate training",
        "teacher model exploitation",
        "model leakage via distillation",
    ],
    "adversarial_ml": [
        "adversarial distillation",
        "defensive distillation",
        "distillation for robustness",
        "robust distillation",
        "adversarial training distillation",
        "transferability via distillation",
        "surrogate adversarial model",
        "attack transfer models",
    ],
    "privacy_leakage": [
        "training data extraction",
        "data leakage machine learning",
        "dataset reconstruction attack",
        "model inversion attack",
        "membership inference attack",
        "attribute inference attack",
        "gradient leakage",
        "information leakage distillation",
        "privacy leakage distillation",
    ],
    "watermarking_ip": [
        "model watermarking",
        "neural network watermarking",
        "distillation resistant watermarking",
        "watermark removal distillation",
        "watermark robustness distillation",
        "model fingerprinting",
        "model ownership verification",
        "IP protection machine learning",
        "model provenance",
        "ownership attacks ML models",
    ],
    "defense_extraction": [
        "model stealing defense",
        "extraction attack defense",
        "query monitoring defense",
        "rate limiting ML APIs",
        "output perturbation defense",
        "confidence masking",
        "prediction obfuscation",
        "defensive noise injection",
        "anti distillation techniques",
        "secure inference systems",
    ],
    "economic_system": [
        "ML as a service attacks",
        "MLaaS security",
        "API abuse machine learning",
        "cloud inference exploitation",
        "cost asymmetry model training vs extraction",
        "compute vs data tradeoff",
        "proprietary model risks",
        "foundation model security",
        "AI system security",
    ],
    "advanced_niche": [
        "distillation collapse",
        "capacity mismatch distillation",
        "student teacher gap",
        "distillation stability",
        "distillation scaling laws",
        "recursive self improvement distillation",
        "iterative distillation",
        "bootstrapped distillation",
        "self training loops",
        "knowledge transfer efficiency",
    ],
    "adjacent_terms": [
        "imitation learning",
        "behavior cloning",
        "surrogate modeling",
        "function approximation copying",
        "black box learning",
        "model compression theory",
        "representation learning transfer",
        "meta learning distillation",
    ],
}


def _phrase_to_regex(phrase: str) -> str:
    escaped = re.escape(phrase.strip())
    flexible = escaped.replace(r"\ ", r"[-\s]+")
    return rf"\b{flexible}\b"

PREFILTER_TERMS = (
    "distill",
    "distil",
    "teacher-student",
    "teacher student",
    "student model",
    "teacher model",
    "model extraction",
    "model stealing",
    "stealing attack",
    "extraction attack",
    "black-box extraction",
    "black box extraction",
    "knockoff",
    "copycat",
    "surrogate model",
    "substitute model",
    "model compression",
    "neural compression",
    "quantization",
    "pruning",
    "data-free",
    "dataset-free",
    "synthetic data",
    "pseudo-label",
    "LLM distillation".casefold(),
    "instruction distillation",
    "alignment distillation",
    "RLHF",
    "preference distillation",
    "reward model",
    "chain-of-thought",
    "reasoning distillation",
    "black-box attack",
    "query-efficient",
    "query synthesis",
    "training data extraction",
    "dataset reconstruction",
    "model inversion",
    "membership inference",
    "gradient leakage",
    "watermark",
    "fingerprint",
    "ownership verification",
    "provenance",
    "extraction defense",
    "query monitoring",
    "rate limiting",
    "output perturbation",
    "prediction obfuscation",
    "confidence masking",
    "anti-distillation",
    "MLaaS",
    "export control",
    "compute asymmetry",
    "supply chain security",
    "capacity gap",
    "scaling law",
)

PREFILTER_TERMS = PREFILTER_TERMS + tuple(
    phrase.casefold()
    for phrases in EXHAUSTIVE_KEYWORD_GROUPS.values()
    for phrase in phrases
)

QUALIFYING_GROUPS = {
    "distillation",
    "distillation_type",
    "compression",
    "data",
    "attack",
    "black_box",
    "leakage",
    "defense",
    "policy",
    "advanced",
    "core_foundational",
    "representation_feature",
    "framework_variant",
    "task_output",
    "llm_foundation",
    "data_free_synthetic",
    "compression_overlap",
    "extraction_stealing",
    "query_black_box",
    "attack_mechanism",
    "adversarial_ml",
    "privacy_leakage",
    "watermarking_ip",
    "defense_extraction",
    "economic_system",
    "advanced_niche",
    "adjacent_terms",
}

WEB_SOURCES = [
    {
        "source_type": "paper",
        "category": "foundation_distillation",
        "title": "Distilling the Knowledge in a Neural Network",
        "url": "https://arxiv.org/abs/1503.02531",
        "notes": "Core Hinton/Vinyals/Dean knowledge distillation paper.",
    },
    {
        "source_type": "paper",
        "category": "model_extraction",
        "title": "Stealing Machine Learning Models via Prediction APIs",
        "url": "https://www.usenix.org/conference/usenixsecurity16/technical-sessions/presentation/tramer",
        "notes": "Canonical prediction API model extraction attack.",
    },
    {
        "source_type": "paper",
        "category": "model_extraction",
        "title": "Knockoff Nets: Stealing Functionality of Black-Box Models",
        "url": "https://arxiv.org/abs/1812.02766",
        "notes": "Black-box functionality stealing via queried input-output pairs.",
    },
    {
        "source_type": "paper",
        "category": "defense",
        "title": "PRADA: Protecting Against DNN Model Stealing Attacks",
        "url": "https://arxiv.org/abs/1805.02628",
        "notes": "Query-distribution monitoring defense for model extraction.",
    },
    {
        "source_type": "paper",
        "category": "data_distillation",
        "title": "Dataset Distillation",
        "url": "https://www.ri.cmu.edu/publications/dataset-distillation/",
        "notes": "Synthetic compact training data as dataset-level distillation.",
    },
    {
        "source_type": "paper",
        "category": "data_free",
        "title": "Dreaming to Distill: Data-free Knowledge Transfer via DeepInversion",
        "url": "https://research.nvidia.com/publication/2020-06_dreaming-distill-data-free-knowledge-transfer-deepinversion",
        "notes": "Data-free knowledge transfer using synthesized images.",
    },
    {
        "source_type": "paper",
        "category": "data_free",
        "title": "Data-Free Knowledge Distillation for Object Detection",
        "url": "https://research.nvidia.com/publication/2021-01_data-free-knowledge-distillation-object-detection",
        "notes": "DeepInversion-style object detection distillation.",
    },
    {
        "source_type": "review",
        "category": "data_distillation",
        "title": "Dataset Distillation: A Comprehensive Review",
        "url": "https://pubmed.ncbi.nlm.nih.gov/37815974/",
        "notes": "TPAMI review of dataset distillation methods.",
    },
    {
        "source_type": "review",
        "category": "watermarking_defense",
        "title": "A Systematic Review on Model Watermarking for Neural Networks",
        "url": "https://www.frontiersin.org/articles/10.3389/fdata.2021.729663/full",
        "notes": "Survey covers watermark robustness, distillation removal, and fingerprinting.",
    },
    {
        "source_type": "paper",
        "category": "model_extraction_monitoring",
        "title": "Model extraction warning in MLaaS paradigm",
        "url": "https://research.ibm.com/publications/model-extraction-warning-in-mlaas-paradigm",
        "notes": "IBM ACSAC work on monitoring extraction status from API query/response streams.",
    },
    {
        "source_type": "paper",
        "category": "llm_reasoning_distillation",
        "title": "Distilling Step-by-Step! Outperforming Larger Language Models with Less Training Data and Smaller Model Sizes",
        "url": "https://arxiv.org/abs/2305.02301",
        "notes": "Reasoning/rationale distillation into smaller language models.",
    },
    {
        "source_type": "paper",
        "category": "llm_reasoning_distillation",
        "title": "Symbolic Chain-of-Thought Distillation: Small Models Can Also Think Step-by-Step",
        "url": "https://arxiv.org/abs/2306.14050",
        "notes": "Chain-of-thought style rationale distillation.",
    },
    {
        "source_type": "paper",
        "category": "llm_instruction_distillation",
        "title": "Distilling Instruction-following Abilities of Large Language Models with Task-aware Curriculum Planning",
        "url": "https://arxiv.org/abs/2405.13448",
        "notes": "Instruction distillation for small LLMs.",
    },
    {
        "source_type": "blog",
        "category": "llm_distillation_product",
        "title": "Model Distillation in the API",
        "url": "https://openai.com/index/api-model-distillation/",
        "notes": "OpenAI product/technical workflow for model distillation.",
    },
    {
        "source_type": "docs",
        "category": "llm_distillation_product",
        "title": "OpenAI distillation guide",
        "url": "https://platform.openai.com/docs/guides/distillation",
        "notes": "Developer workflow for distilling larger model outputs into smaller fine-tuned models.",
    },
    {
        "source_type": "technical_discussion",
        "category": "policy_economics",
        "title": "How to Steal an AI",
        "url": "https://www.wired.com/2016/09/how-to-steal-an-ai",
        "notes": "Accessible discussion of prediction API extraction and business/security implications.",
    },
    {
        "source_type": "news",
        "category": "policy_economics",
        "title": "Anthropic says DeepSeek and other Chinese AI companies fraudulently used Claude",
        "url": "https://www.businessinsider.com/anthropic-deepseek-distillation-minimax-moonshot-ai-2026-2",
        "notes": "Recent policy/economic framing of distillation and model extraction allegations.",
    },
    {
        "source_type": "news",
        "category": "policy_economics",
        "title": "Anthropic accuses Chinese AI labs of mining Claude as US debates AI chip exports",
        "url": "https://techcrunch.com/2026/02/23/anthropic-accuses-chinese-ai-labs-of-mining-claude-as-us-debates-ai-chip-exports/",
        "notes": "Distillation tied to export controls and frontier model competition.",
    },
]


def _compiled_patterns() -> list[tuple[str, str, re.Pattern[str], int]]:
    compiled = [
        (group, label, re.compile(pattern, re.IGNORECASE), weight)
        for group, entries in PATTERN_GROUPS.items()
        for label, pattern, weight in entries
    ]
    compiled.extend(
        (
            group,
            phrase,
            re.compile(_phrase_to_regex(phrase), re.IGNORECASE),
            4,
        )
        for group, phrases in EXHAUSTIVE_KEYWORD_GROUPS.items()
        for phrase in phrases
    )
    return compiled


PATTERNS = _compiled_patterns()


def _paper_source(conf_dir: Path) -> Path | None:
    enriched = conf_dir / "papers_enriched.jsonl"
    plain = conf_dir / "papers.jsonl"
    if enriched.exists() and enriched.stat().st_size > 0:
        return enriched
    if plain.exists() and plain.stat().st_size > 0:
        return plain
    return None


def _load_papers(data_dir: Path) -> Iterable[tuple[str, dict]]:
    for conf_dir in sorted(data_dir.iterdir()):
        if not conf_dir.is_dir() or conf_dir.name.startswith("."):
            continue

        source = _paper_source(conf_dir)
        if source is None:
            continue

        with open(source, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                yield conf_dir.name, json.loads(line)


def _load_master_csv(path: Path) -> Iterable[tuple[str, dict]]:
    if not path.exists() or path.stat().st_size == 0:
        return

    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = row.get("title", "")
            if not title:
                continue

            venue = (row.get("venue") or "master_literature").strip() or "master_literature"
            year = (row.get("year") or "").strip()
            conf_id = f"{venue.lower().replace(' ', '_')}_{year}" if year else venue.lower().replace(" ", "_")
            yield conf_id, {
                "title": title,
                "authors": [a.strip() for a in (row.get("authors") or "").split(",") if a.strip()],
                "abstract": row.get("abstract", ""),
                "link": row.get("url", ""),
                "citation_count": row.get("citations") or 0,
                "source": row.get("source", "master_literature"),
            }


def _field_text(paper: dict) -> dict[str, str]:
    keywords = paper.get("keywords", [])
    if isinstance(keywords, list):
        keywords_text = " ".join(str(k) for k in keywords)
    else:
        keywords_text = str(keywords or "")

    return {
        "title": str(paper.get("title", "")),
        "abstract": str(paper.get("abstract", "")),
        "tldr": str(paper.get("tldr", "")),
        "keywords": keywords_text,
    }


def _score_paper(paper: dict) -> tuple[int, list[str], list[str]]:
    fields = _field_text(paper)
    combined = " ".join(fields.values()).casefold()
    if not any(term in combined for term in PREFILTER_TERMS):
        return 0, [], []

    matched_terms: dict[str, set[str]] = {}
    matched_groups: set[str] = set()
    score = 0

    for field_name, text in fields.items():
        if not text:
            continue
        field_weight = FIELD_WEIGHTS[field_name]

        for group, label, pattern, weight in PATTERNS:
            if not pattern.search(text):
                continue
            matched_groups.add(group)
            matched_terms.setdefault(label, set()).add(field_name)
            score += field_weight * weight

    if not (matched_groups & QUALIFYING_GROUPS):
        return 0, [], []

    if "llm" in matched_groups and "distillation" in matched_groups:
        score += 15
    if "attack" in matched_groups and "distillation" in matched_groups:
        score += 12
    if "black_box" in matched_groups and "attack" in matched_groups:
        score += 10
    if "defense" in matched_groups and "attack" in matched_groups:
        score += 8
    if "data" in matched_groups and ("distillation" in matched_groups or "attack" in matched_groups):
        score += 8
    if "policy" in matched_groups and ("attack" in matched_groups or "defense" in matched_groups):
        score += 6

    terms = [
        f"{term} ({', '.join(sorted(fields))})"
        for term, fields in sorted(matched_terms.items())
    ]
    return score, sorted(matched_groups), terms


def _parse_conf_id(conf_id: str) -> tuple[str, str]:
    venue, sep, year = conf_id.rpartition("_")
    if sep and year.isdigit():
        return venue, year
    return conf_id, ""


def _dedupe_key(paper: dict) -> str:
    title = re.sub(r"\s+", " ", str(paper.get("title", "")).casefold()).strip()
    authors = paper.get("authors", [])
    first_author = str(authors[0]).casefold() if authors else ""
    return f"{title}|{first_author}"


def _citation_sort_value(paper: dict) -> int:
    citations = paper.get("citation_count")
    if citations is None:
        citations = paper.get("citations")
    try:
        return int(citations or 0)
    except (TypeError, ValueError):
        return 0


def extract(data_dir: Path, min_score: int, master_csv: Path | None = None) -> tuple[list[dict], int]:
    seen: dict[str, dict] = {}
    scanned = 0

    sources = [_load_papers(data_dir)]
    if master_csv is not None:
        sources.append(_load_master_csv(master_csv))

    for conf_id, paper in (item for source in sources for item in source):
        scanned += 1
        score, groups, terms = _score_paper(paper)
        if score < min_score:
            continue

        venue, year = _parse_conf_id(conf_id)
        if not year:
            venue = str(paper.get("source_name") or venue)
        authors = paper.get("authors", [])
        if isinstance(authors, list):
            authors_text = "; ".join(str(a) for a in authors)
        else:
            authors_text = str(authors or "")

        result = {
            "score": score,
            "groups": groups,
            "matched_terms": terms,
            "conference_id": conf_id,
            "venue": venue,
            "year": year,
            "title": paper.get("title", ""),
            "authors": authors_text,
            "link": paper.get("link") or paper.get("url", ""),
            "citation_count": _citation_sort_value(paper),
            "tldr": paper.get("tldr", ""),
            "abstract": paper.get("abstract", ""),
        }

        key = _dedupe_key(paper)
        existing = seen.get(key)
        if existing is None or (score, result["citation_count"]) > (existing["score"], existing["citation_count"]):
            seen[key] = result

    results = sorted(
        seen.values(),
        key=lambda p: (p["score"], p["citation_count"], p["year"]),
        reverse=True,
    )
    return results, scanned


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def _flatten_candidate(row: dict) -> dict:
    return {
        **row,
        "groups": "; ".join(row["groups"]),
        "matched_terms": "; ".join(row["matched_terms"]),
    }


def _write_csv(path: Path, rows: list[dict]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow(_flatten_candidate(row))


def _write_web_sources_csv(path: Path) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=WEB_SOURCE_FIELDNAMES)
        writer.writeheader()
        writer.writerows(WEB_SOURCES)


def _write_report(path: Path, rows: list[dict], scanned: int) -> None:
    by_group = Counter(group for row in rows for group in row["groups"])
    by_venue = Counter(row["venue"] for row in rows)
    by_year = Counter(row["year"] for row in rows if row["year"])

    lines = [
        "# Distillation Candidate Papers",
        "",
        f"- Papers scanned: {scanned}",
        f"- Candidate papers: {len(rows)}",
        f"- Precision note: title-only DBLP rows are included when their title matches; enrich those venues for better recall.",
        "",
        "## Group Counts",
        "",
    ]
    lines.extend(f"- {group}: {count}" for group, count in by_group.most_common())
    lines.extend(["", "## Top Venues", ""])
    lines.extend(f"- {venue}: {count}" for venue, count in by_venue.most_common(20))
    lines.extend(["", "## Top Years", ""])
    lines.extend(f"- {year}: {count}" for year, count in by_year.most_common(20))
    lines.extend(["", "## Top Candidates", ""])

    for row in rows[:100]:
        citations = row["citation_count"]
        citation_text = f", {citations} citations" if citations else ""
        groups = ", ".join(row["groups"])
        terms = "; ".join(row["matched_terms"][:8])
        lines.append(f"- **{row['title']}** ({row['conference_id']}, score {row['score']}{citation_text})")
        lines.append(f"  Groups: {groups}")
        lines.append(f"  Matches: {terms}")
        if row["link"]:
            lines.append(f"  Link: {row['link']}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_subset_outputs(output_prefix: Path, rows: list[dict]) -> None:
    subsets = {
        "llm": [row for row in rows if "llm" in row["groups"]],
        "non_llm": [row for row in rows if "llm" not in row["groups"]],
        "attacks": [row for row in rows if "attack" in row["groups"]],
        "black_box": [row for row in rows if "black_box" in row["groups"]],
        "defenses": [row for row in rows if "defense" in row["groups"]],
        "data_synthetic": [row for row in rows if "data" in row["groups"]],
        "compression": [row for row in rows if "compression" in row["groups"]],
        "policy": [row for row in rows if "policy" in row["groups"]],
        "techniques": [row for row in rows if "technique" in row["groups"]],
    }

    for name, subset in subsets.items():
        subset_prefix = output_prefix.with_name(f"{output_prefix.name}_{name}")
        _write_jsonl(subset_prefix.with_suffix(".jsonl"), subset)
        _write_csv(subset_prefix.with_suffix(".csv"), subset)


def _excel_value(value):
    if isinstance(value, list):
        value = "; ".join(str(v) for v in value)
    text = "" if value is None else str(value)
    if len(text) > EXCEL_CELL_LIMIT:
        return text[: EXCEL_CELL_LIMIT - 3] + "..."
    return text


def _append_sheet(workbook, title: str, rows: list[dict], fieldnames: list[str]) -> None:
    sheet = workbook.create_sheet(title[:31])
    sheet.append(fieldnames)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        sheet.append([_excel_value(row.get(field, "")) for field in fieldnames])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        column_letter = column[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column[:100])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def _taxonomy_rows() -> list[dict]:
    rows = []
    for group, entries in PATTERN_GROUPS.items():
        for label, pattern, weight in entries:
            rows.append({
                "group": group,
                "term": label,
                "regex": pattern,
                "weight": weight,
            })
    return rows


def _write_xlsx(path: Path, rows: list[dict]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    subsets = {
        "all_candidates": rows,
        "llm": [row for row in rows if "llm" in row["groups"]],
        "non_llm": [row for row in rows if "llm" not in row["groups"]],
        "attacks": [row for row in rows if "attack" in row["groups"]],
        "black_box": [row for row in rows if "black_box" in row["groups"]],
        "defenses": [row for row in rows if "defense" in row["groups"]],
        "data_synthetic": [row for row in rows if "data" in row["groups"]],
        "compression": [row for row in rows if "compression" in row["groups"]],
        "policy": [row for row in rows if "policy" in row["groups"]],
        "techniques": [row for row in rows if "technique" in row["groups"]],
    }

    for sheet_name, subset in subsets.items():
        _append_sheet(
            workbook,
            sheet_name,
            [_flatten_candidate(row) for row in subset],
            FIELDNAMES,
        )

    _append_sheet(workbook, "web_sources", WEB_SOURCES, WEB_SOURCE_FIELDNAMES)
    _append_sheet(workbook, "query_taxonomy", _taxonomy_rows(), ["group", "term", "regex", "weight"])
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract distillation-related papers from local data.")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--master-csv", type=Path, default=Path("master_literature.csv"))
    parser.add_argument("--no-master-csv", action="store_true", help="Skip master_literature.csv even if present.")
    parser.add_argument("--output-prefix", type=Path, default=DEFAULT_OUTPUT_PREFIX)
    parser.add_argument("--min-score", type=int, default=12)
    parser.add_argument("--limit", type=int, default=0, help="Optional max rows to write after ranking.")
    args = parser.parse_args()

    master_csv = None if args.no_master_csv else args.master_csv
    rows, scanned = extract(args.data_dir, min_score=args.min_score, master_csv=master_csv)
    if args.limit:
        rows = rows[: args.limit]

    args.output_prefix.parent.mkdir(parents=True, exist_ok=True)
    _write_jsonl(args.output_prefix.with_suffix(".jsonl"), rows)
    _write_csv(args.output_prefix.with_suffix(".csv"), rows)
    _write_web_sources_csv(args.output_prefix.with_name(f"{args.output_prefix.name}_web_sources").with_suffix(".csv"))
    _write_report(args.output_prefix.with_suffix(".md"), rows, scanned)
    _write_subset_outputs(args.output_prefix, rows)
    _write_xlsx(args.output_prefix.with_suffix(".xlsx"), rows)

    print(f"Scanned {scanned} papers")
    print(f"Wrote {len(rows)} candidates to {args.output_prefix.with_suffix('.csv')}")
    print(f"Excel workbook: {args.output_prefix.with_suffix('.xlsx')}")
    print(f"Report: {args.output_prefix.with_suffix('.md')}")


if __name__ == "__main__":
    main()
